import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook() #creates workbook object

ws1 = wb.active # essentially the first sheet

ws1.title = 'First Sheet'

ws2 = wb.create_sheet(index = 1, title = 'Second Sheet') #creatinf the second sheet

ws1['A1'] = 'Invoice'

myfont = Font(name = 'Times New Roman', size =24, bold = True, italic = False)

ws1['A1'].font = myfont

ws1['A2'] =  'Tires'
ws1['A3'] = 'Breaks'
ws1['A4'] = 'Alignment'

ws1.merge_cells = ['A1:B1'] #Merges the cells

ws1['B2'] =  450
ws1['B3'] =  225
ws1['B4'] =  150

ws1['A8'] = 'Total'
ws1['A8'].font = myfont

ws1['B8'] = '=SUM(B2:B4)' # copies from excel

#ws.column_dimensions['A'] = 25




#Read the exxcel file 'ProduceReport.xlsx' that you created earlier
#Write all the contents of this file to 'Second Sheet' in the current workbook

#Display the Grand Total and Average of 'Amn Sold' and 'Total'
#At the bottom of the list along with appropriate labels

write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row
print(maxC)
print(maxR)

write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost Per Pound'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'

write_row = 2 # the row we wanna start write to
write_colA = 1 #show what columns your working with
write_colB = 2
write_colC = 3
write_colD = 4

for currentrow in read_ws.iter_rows(min_row = 2,max_row = maxR, max_col = maxC):
    
    #go through each row of the read and set cell values as variables
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = float(currentrow[2].value)
    total = float(currentrow[3].value)

    #Sets the values of each cell according to the variables we set in the reading part
    write_sheet.cell(write_row,write_colA).value = name 
    write_sheet.cell(write_row,write_colB).value = cost
    write_sheet.cell(write_row,write_colC).value = amt_sold
    write_sheet.cell(write_row,write_colD).value = total

    write_row += 1 # do this to make it goto the next wrow

summary_row = write_row + 1

write_sheet['B' + str(summary_row)] = 'Total'
write_sheet['C' + str(summary_row)] = '=SUM(C2:C' + str(write_row) + ')'
write_sheet['D' + str(summary_row)] = '=SUM(D2:D' + str(write_row) + ')'

summary_row += 1

write_sheet['B' + str(summary_row)] = 'Average'
write_sheet['C' + str(summary_row)] = '=AVERAGE(C2:C' + str(write_row) + ')'
write_sheet['D' + str(summary_row)] = '=AVERAGE(D2:D' + str(write_row) + ')'

for cell in write_sheet["C:C"]: #chnage the format for each cell
    cell.number_format = '#,##0'
    
for cell in write_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'





""" ws2['D42'] = 'Grand Total'
ws2['D43'] = '=SUM(D2:D41)'

ws2['C42'] = 'Average Sold'    
ws2['C43'] = '=AVERAGE(C2:C41)' """
        #print(ws2['A2'].value)







wb.save('PythontoExcel.xlsx')
