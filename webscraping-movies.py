
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

#scrape top 5 movies with number,name, release date, total gross, % of total gross



#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2023/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

wb = xl.Workbook()
ws = wb.active

ws.ttile = 'Box Office Report'

ws['A1'] = 'No. '
ws['B1'] = 'Tile'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = '% of Total Gross'


title = soup.title

print(title.text)
print()

movie_rows = soup.find_all("tr")

for x in range(1,6): #dont wanna start at row 0/ first tr tag cuz it has nothing
    td = movie_rows[x].find_all('td') 
    print(td[1].text)
    number = td[0].text
    title = td[1].text
    gross = int(td[5].text.replace(",", "").replace("$", ""))
    total_gross = int(td[7].text.replace(",", "").replace("$", ""))
    release_date = td[8].text

    percent_gross = gross / total_gross

    #now staret writing to excell

    ws['A' + str(x + 1)] = number #starts on the second row
    ws['B' + str(x + 1)] = title
    ws['C' + str(x + 1)] = release_date
    ws['D' + str(x + 1)] = gross
    ws['E' + str(x + 1)] = total_gross
    ws['F' + str(x + 1)] = str(percent_gross) + "%" # convert to string to concantenate

ws.column_dimensions['A'].width = 5

header_font = Font(size=16, bold = True)

for cell in ws[1:1]:
    cell.font = header_font

wb.save("BoxOfficeReport.xlsx")
#print(table_rows)
# class might only need to be specified if their are multiple tables on the page
# ^ you already find the row, so now drill to eavh cell from the row


