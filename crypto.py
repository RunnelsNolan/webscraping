from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font, PatternFill



webpage = 'https://sahicoin.com/en-us/cryptocurrencies-top-gainer'
page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

wb = xl.Workbook()
ws = wb.active
ws.title = "Top Gainers"

ws['A1'] = 'Crypto Name'
ws['B1'] = 'Price'
ws['C1'] = '% Change'
ws['D1'] = 'Price Change'
ws['E1'] = 'Original Price'



title = soup.title
print(title.text)

header_fill = PatternFill(patternType= 'solid', fgColor= 'FFB266')
header_font = Font(size=16, bold= True)

for cell in ws[1:1]:
    cell.font = header_font
    cell.fill = header_fill
name_font = Font(size = 11, italic=True)
percent_font = Font(size = 11, color= "008000")

crypto_rows = soup.find_all("tr")
for x in range(1,6):
    td = crypto_rows[x].find_all('td')
    name = td[1].text
    price = float(td[2].text.replace("$",""))
    change_percent = float(td[3].text.replace("%","").replace(",",""))
    
    true_percent = change_percent / 100
    original_price = price / (1+true_percent)
    price_change = price - original_price

    ws['A' + str(x + 1)] = name
    ws['A' + str(x + 1)].font = name_font

    ws['B' + str(x + 1)] = '$' + str(format(price, ',.2f'))

    ws['C' + str(x + 1)] = str(change_percent) + '%'
    ws['C' + str(x + 1)].font = percent_font

    ws['D' + str( x + 1)] = '$' + str(format(price_change, ',.4f')) 
    ws['E' + str(x + 1)] = '$' + str(format(original_price, ',.4f') ) 

ws.column_dimensions['A'].width = 23
ws.column_dimensions['B'].width = 13
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 16

#header_font = Font(size=16, bold= True)







 


wb.save("Crypto_Gainers.xlsx")


#Bitcoin stuff

bit_webpage = 'https://sahicoin.com/en-us/cryptocurrency-bitcoin'

bit_page = urlopen(bit_webpage)			

soup = BeautifulSoup(bit_page, 'html.parser')


bitCoin = soup.find('h2', class_='price')
bitCoin_price = float(bitCoin.text.replace("$","").replace(",", ""))


bitChange = soup.find('span', class_='price_status')
bitChange = float(bitChange.text.replace("%", "")) /100


bit_original_price = bitCoin_price / (1+bitChange)
bit_price_change = bitCoin_price - bit_original_price
#print(bit_price_change)



#Ethereum Data

eth_webpage = 'https://sahicoin.com/en-us/cryptocurrency-ethereum'

eth_page = urlopen(eth_webpage)			

soup = BeautifulSoup(eth_page, 'html.parser')


ethereum = soup.find('h2', class_='price')
ethereum_price = float(ethereum.text.replace("$","").replace(",", ""))



ethChange = soup.find('span', class_='price_status')
ethChange = float(ethChange.text.replace("%", "")) /100


eth_original_price = ethereum_price / (1+ethChange)
eth_price_change = ethereum_price - eth_original_price
#print(eth_price_change)






import keys 
from twilio.rest import Client

client = Client(keys.accountSID, keys.auth_token)

TwilioNumber = "+15074605741"

mycellphone = "+12145194721" 


#TEXT MESSAGE STUFF


if bit_price_change >= 5 or bit_price_change <= 5:
    textmessage = client.messages.create(to=mycellphone, from_= TwilioNumber,body= f'Bit Coin has chnaged by ${bit_price_change} ')
    print(textmessage.status)

if eth_price_change >= 5 or eth_price_change <= 5:
    textmessage = client.messages.create(to=mycellphone, from_= TwilioNumber,body= f'Ethereum has chnaged by ${eth_price_change} ')
    print(textmessage.status)
