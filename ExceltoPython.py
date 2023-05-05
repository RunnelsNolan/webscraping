import openpyxl as xl

wb = xl.load_workbook('example.xlsx') # opens the file up
sn = wb.sheetnames # gives sheet names
print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(cellA1.value) # gives data in the cell
print(type(cellA1.value)) #enters python as same data type as in excel
print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate) # guves row and column

print(sheet1.cell(1,2).value) # give it the row and column number to specify a cell
#starts at 1 and 1 not index value of zero

print(sheet1.max_row) #gives last row with data in it
print(sheet1.max_column) 

for i in range(1,sheet1.max_row + 1): #give name of all fruits , RANGE is not inclusive
    print(sheet1.cell(i,2).value)

print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(900))

#print(xl.utils.get_column_index_from_string('AHP')) # if you know column letter that u wanna goto, use this to get its number

for currentrow in sheet1['A1': 'C3']:
    print(currentrow) #give each row
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)

for currentrow in sheet1.iter_rows(min_row = 2, max_row= 7, max_col= sheet1.max_column):
    print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)


