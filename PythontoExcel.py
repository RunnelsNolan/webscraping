import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook() #creates workbook object

ws1 = wb.active # essentially the first sheet

ws1.title = 'First Sheet'

ws2 = wb.create_sheet(index = 1, title = 'Second Sheet') #creatinf the second sheet

ws1['A1'] = 'Invoice'

myfont = Font(name = 'Times New Roman', size =24, bold = True, italic = False)

ws1['A1'].font = myfont

ws1['A2'] =  'Tires'
ws1['A3'] = 'Breaks'
ws1['A4'] = 'Alignment'

ws1.merge_cells = ['A1:B1'] #Merges the cells

ws1['B2'] =  450
ws1['B3'] =  225
ws1['B4'] =  150

ws1['A8'] = 'Total'
ws1['A8'].font = myfont

ws1['B8'] = '=SUM(B2:B4)' # copies from excel

#ws.column_dimensions['A'] = 25




#Read the exxcel file 'ProduceReport.xlsx' that you created earlier
#Write all the contents of this file to 'Second Sheet' in the current workbook

#Display the Grand Total and Average of 'Amn Sold' and 'Total'
#At the bottom of the list along with appropriate labels

Prod_wb = xl.load_workbook('ProduceReport.xlsx')
sn = Prod_wb.sheetnames
print(sn)

Prod_sheet1 = Prod_wb['ProduceReport']

for currentrow in  Prod_sheet1.iter_rows(min_row=1,max_row =Prod_sheet1.max_row, max_col= Prod_sheet1.max_column):
    for currentcell in currentrow:
        coordinate = str(currentcell.coordinate)
        
        ws2[coordinate] = currentcell.value

ws2['D42'] = 'Grand Total'
ws2['D43'] = '=SUM(D2:D41)'

ws2['C42'] = 'Average Sold'    
ws2['C43'] = '=AVERAGE(C2:C41)'
        #print(ws2['A2'].value)







wb.save('PythontoExcel.xlsx')
