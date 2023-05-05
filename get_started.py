from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

webpage = 'https://registrar.web.baylor.edu/exams-grading/spring-2023-final-exam-schedule'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

wb = xl.Workbook()
ws = wb.active




ws['A1'] = ''
ws['B1'] = ''
ws['C1'] = ''
ws['D1'] = ''

    #ws['A' + str(x + 1)] = number
    #ws['B' + str(x + 1)] = country
    #ws['C' + str(x + 1)] = gdp
    #ws['D' + str(x + 1)] = population


wb.save(".xlsx")